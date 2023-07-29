import openpyxl

from alumni_info import Alumni
from alumni_manage_services import Alumni_Management_Services

import pymysql
connection = pymysql.connect(user='root',password='root',db='alumnidb',host='localhost',port=3306)

class Alumni_Services_Impl(Alumni_Management_Services):
    alumni_list = []

    def add_alumni(self,alumni):
        Alumni_Services_Impl.alumni_list.append(alumni)
        try:
            channel = connection.cursor()
            INSERT_ALUMNI_DATA = f'''
                INSERT INTO ALUMNI_MASTER VALUES({alumni.alumni_id},'{alumni.alumni_name}',
                '{alumni.alumni_department}','{alumni.alumni_company}',
                {alumni.year_of_graduation},{alumni.alumni_number})
                '''
            channel.execute(INSERT_ALUMNI_DATA)
        except BaseException as e:
            print('Problem in Alumni Add ! ',e.args)
        else:
            connection.commit()
            print('Alumni added successfully! ')
        finally:
            channel.close()

    def delete_alumni(self,alumniid):
        try:
            channel = connection.cursor()
            DELETE_ALUMNI = f'''DELETE FROM ALUMNI_MASTER WHERE ALUMNI_ID = {alumniid}; '''
            channel.execute(DELETE_ALUMNI)
        except BaseException as e:
            print("problem in delete alumni! ",e.args)
        else:
            connection.commit()
            print("alumni deleted successfully !")
        finally:
            channel.close()

    def search_alumni(self,criteria,value):
        try:
            channel = connection.cursor()
            if criteria == 'ID':
                search = f'''
                SELECT * FROM ALUMNI_MASTER WHERE ALUMNI_ID = '{value}'
                '''
            elif criteria == 'NAME':
                search = f'''
                SELECT * FROM ALUMNI_MASTER WHERE ALUMNI_NAME = '{value}'
                '''
            elif criteria == 'DEPARTMENT':
                search = f'''
                SELECT * FROM ALUMNI_MASTER WHERE ALUMNI_DEPARTMENT = '{value}'
                '''

            elif criteria == 'COMPANY':
                search = f'''
                SELECT * FROM ALUMNI_MASTER WHERE ALUMNI_COMPANY = '{value}'
                '''

            elif criteria == 'YEAR_OF_GRADUATION':
                search = f'''
                SELECT * FROM ALUMNI_MASTER WHERE YEAR_OF_GRADUATION = '{value}'
                '''
            channel.execute(search)
            rows = channel.fetchall()
        except BaseException as e:
            print("Problem in search ! ",e.args)
        else:
            print("alimni successfully searched! ")
        finally:
            channel.close()
        return rows

    def list_alumni(self,criteria = None,value = None):
        try:
            List_Alumni = '''SELECT * FROM ALUMNI_MASTER'''
            channel = connection.cursor()
            channel.execute(List_Alumni)
            results = channel.fetchall()
            if results:
                for row in results:
                    print(row)
            else:
                print("No alumni found.")
        except BaseException as e:
            print('problem in listing!',e.args)
        finally:
            channel.close()

    def alumni_of_year_in_range(self,from_year,to_year):
        try:
            IN_YEAR = f'''
            SELECT * FROM alumni_master
            WHERE YEAR_OF_GRADUATION BETWEEN {from_year} AND {to_year};'''
            channel = connection.cursor()
            channel.execute(IN_YEAR)
            results = channel.fetchall()
            if results:
                for row in results:
                    print(row)
            else:
                print("No alumni found within the specified year range.")
        except BaseException as e:
            print("problem in range function!..",e.args)
        finally:
            channel.close()



    def export_data_excel(self):
        import openpyxl
        try:
            query = "SELECT * FROM ALUMNI_MASTER"
            channel = connection.cursor()
            channel.execute(query)
            data = channel.fetchall()
            file_path = input("Enter excel file name with .xlsx extention!: ")


            workbook = openpyxl.Workbook()
            sheet1 = workbook.create_sheet('Alumni_Data-1')
            # sheet2 = workbook.create_sheet('Alumni_Data-2')
            # sheet3 = workbook.create_sheet('Alumni_Data-3')

            # headers
            sheet1['A1'] = 'ALUMNI_ID'
            sheet1['B1'] = 'ALUMNI_NAME'
            sheet1['C1'] = 'ALUMNI_DEPARTMENT'
            sheet1['D1'] = 'ALUMNI_COMPANY'
            sheet1['E1'] = 'YEAR_OF_GRADUATION'
            sheet1['F1'] = 'ALUMNI_NUMBER'
            #print("Headers are Created successfully! ")

            row_number = 2
            for row_data in data:
                sheet1[f"A{row_number}"] = row_data[0]  # ALUMNI_ID
                sheet1[f"B{row_number}"] = row_data[1]  # ALUMNI_NAME
                sheet1[f"C{row_number}"] = row_data[2]  # ALUMNI_DEPARTMENT
                sheet1[f"D{row_number}"] = row_data[3]  # ALUMNI_COMPANY
                sheet1[f"E{row_number}"] = row_data[4]  # YEAR OF GRADUATION
                sheet1[f"F{row_number}"] = row_data[5]  # ALUMNI_NUMBER
                row_number = row_number + 1
            workbook.save(file_path)
            print("Excel file is created successfully!")
        except BaseException as e:
            print("problem in creating excel file! ",e.args)
        finally:
            workbook.close()
            channel.close()


    def import_excel_data(self):
        file_path = input("Enter the name of excel file for importing!: ")
        try:
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook['Alumni_Data-1']
            data = []
            for row in sheet.iter_rows(min_row=2,values_only=True):
                data.append(row)
            query = "INSERT INTO ALUMNI_MASTER (ALUMNI_ID, ALUMNI_NAME, ALUMNI_DEPARTMENT, ALUMNI_COMPANY, YEAR_OF_GRADUATION, ALUMNI_NUMBER) VALUES (%s, %s, %s, %s, %s, %s)"

            channel = connection.cursor()
            channel.executemany(query,data)
            connection.commit()
            print("Data imported from Excel file successfully!")
        except BaseException as e:
            print('problem in importing excel file!',e.args)
        finally:
            channel.close()







